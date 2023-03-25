import requests
import bs4
import os
import json

from pymongo import MongoClient as MC

# For Excel Files
import openpyxl 

uri = "mongodb://localhost:27017"

client = MC(uri)

db = client['fundtrail']
col = db['m2']

def inDb(data):
    col.insert_many(data)
    print("Data Inseted to Database Successfully")



def xlToData():
    wb = openpyxl.load_workbook('data/miniBank.xlsx')
    ws = wb.active
    print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))
    

    m = []
    # for k in range(ws.max_rows)
    for i in range(2,ws.max_row):
        final = {}
        for j in range(1,ws.max_column):
            final[str(ws.cell(row=1, column=j).value)] = ws.cell(row=i, column=j).value

        # data is finilised here
        # print(final)
        m.append(final)
    inDb(m)




xlToData()


